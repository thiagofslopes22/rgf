"""
conciliador_rreo.py
────────────────────────────────────────────────────────────────────────
Motor de conciliação do RREO (Relatório Resumido de Execução Orçamentária).

Recebe dois arquivos (.xls legado ou .xlsx) — o rascunho gerado pela MSC e
o arquivo homologado no SICONFI — e gera um .xlsx que é cópia fiel do
layout/estrutura do rascunho, com as células divergentes destacadas
(cor + comentário com valor MSC, valor SICONFI, diferença e classificação).

Mesma filosofia do conciliador.py do RGF: nenhuma camada de IA na
comparação. É leitura estruturada (xlrd, preserva fonte/cor/merge/largura)
+ diff célula-a-célula. 100% determinístico e auditável — requisito básico
pra um produto que vai ser usado por auditores/TCE.

O RREO tem layout multi-anexo (várias abas: Anexo 01, 02, 03, 04, 06, 07,
13, 14...) e cada aba tem cabeçalho composto (2-3 linhas de título de
coluna) em vez do cabeçalho de uma linha do RGF. A regra de "zero vs
vazio" também é diferente: no SICONFI homologado, campos sem movimento
aparecem como 0,00; no rascunho da MSC, o mesmo campo pode vir vazio.
Isso NÃO é divergência — é só forma de representar "nada a reportar".

Classificação de divergência (mesma filosofia do RGF, adaptada a contexto
fiscal — RREO compara valores absolutos em R$, então a severidade é
baseada no percentual de variação sobre o valor homologado):

  CRÍTICA         diff percentual > 20%
  SIGNIFICATIVA   5% – 20%
  MODERADA        1% – 5%
  BAIXA           < 1%
  MÍNIMA          diferença absoluta < R$ 0,01 (arredondamento)
  AUSENTE_MSC     valor presente no SICONFI, ausente/vazio no rascunho MSC
  AUSENTE_SICONFI valor presente no rascunho MSC, ausente/vazio no SICONFI
  TEXTO           célula de texto (não numérica) divergente, ex. notas de
                  publicação no rodapé do anexo

Uso:
    from conciliador_rreo import conciliar_rreo
    resultado = conciliar_rreo("rascunho.xls", "siconfi.xls", "saida.xlsx")
    # resultado é um dict com o resumo (ver RESULTADO_SCHEMA no fim do arquivo)
"""

from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
import uuid
from dataclasses import dataclass, field
from typing import Any

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Paleta de severidade (mesmos princípios visuais do RGF) ───────────
DIV_FILLS: dict[str, PatternFill] = {
    "CRÍTICA":          PatternFill("solid", fgColor="FFFF3333"),
    "SIGNIFICATIVA":    PatternFill("solid", fgColor="FFFFAA00"),
    "MODERADA":         PatternFill("solid", fgColor="FFFFFF44"),
    "BAIXA":            PatternFill("solid", fgColor="FF90EE90"),
    "MÍNIMA":           PatternFill("solid", fgColor="FFD4F1D4"),
    "AUSENTE_MSC":      PatternFill("solid", fgColor="FFADD8E6"),
    "AUSENTE_SICONFI":  PatternFill("solid", fgColor="FFC9A0DC"),
    "TEXTO":            PatternFill("solid", fgColor="FFD9D9D9"),
}

# Limiares percentuais de severidade (sobre o valor homologado |SICONFI|)
LIMIAR_CRITICA = 0.20
LIMIAR_SIGNIFICATIVA = 0.05
LIMIAR_MODERADA = 0.01
LIMIAR_MINIMA_ABS = 0.01  # R$ 0,01 — diferença de centavos / arredondamento


# ── Identidade visual Kora (relatório tabular de auditoria) ────────────
KORA_GRAFITE = "FF0E1117"
KORA_DOURADO = "FFC9962F"
KORA_DOURADO_CLARO = "FFE8C77A"
KORA_BRANCO = "FFFFFFFF"
KORA_MARFIM = "FFF7F5F0"
KORA_CINZA_TEXTO = "FF5A5F6B"
KORA_CINZA_BORDA = "FFE2DFD6"

KORA_FONT_FAMILY = "Arial"

CLASS_BADGE_FILLS: dict[str, str] = {
    "CRÍTICA":          "FFF8D0D0",
    "SIGNIFICATIVA":    "FFFCE3BF",
    "MODERADA":         "FFFCF3C0",
    "BAIXA":            "FFDCF0DC",
    "MÍNIMA":           "FFE8F5E8",
    "AUSENTE_MSC":      "FFD9E8F5",
    "AUSENTE_SICONFI":  "FFE9DCF0",
    "TEXTO":            "FFEAEAEA",
}
CLASS_BADGE_TEXT: dict[str, str] = {
    "CRÍTICA":          "FF8A1F1F",
    "SIGNIFICATIVA":    "FF8A5A12",
    "MODERADA":         "FF7A6A10",
    "BAIXA":            "FF1F6B2E",
    "MÍNIMA":           "FF2E7D3A",
    "AUSENTE_MSC":      "FF1F4E78",
    "AUSENTE_SICONFI":  "FF5A2E78",
    "TEXTO":            "FF555555",
}


# ── Estruturas de retorno ──────────────────────────────────────────────
@dataclass
class Divergencia:
    sheet: str
    row: int
    col: str
    col_idx: int
    rotulo: str | None
    cabecalho: str | None
    valor_msc: Any
    valor_siconfi: Any
    diferenca: float | None
    diferenca_pct: float | None
    classificacao: str

    def to_dict(self) -> dict:
        return {
            "sheet": self.sheet,
            "celula": f"{self.col}{self.row}",
            "rotulo": self.rotulo,
            "cabecalho": self.cabecalho,
            "valor_msc": self.valor_msc,
            "valor_siconfi": self.valor_siconfi,
            "diferenca": self.diferenca,
            "diferenca_pct": self.diferenca_pct,
            "classificacao": self.classificacao,
        }


@dataclass
class ResultadoConciliacao:
    arquivo_saida: str
    arquivo_auditoria: str | None
    total_divergencias: int
    divergencias: list[Divergencia] = field(default_factory=list)
    por_sheet: dict[str, int] = field(default_factory=dict)
    por_classificacao: dict[str, int] = field(default_factory=dict)
    sheets_comparadas: list[str] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "arquivo_saida": self.arquivo_saida,
            "arquivo_auditoria": self.arquivo_auditoria,
            "total_divergencias": self.total_divergencias,
            "por_sheet": self.por_sheet,
            "por_classificacao": self.por_classificacao,
            "sheets_comparadas": self.sheets_comparadas,
            "avisos": self.avisos,
            "divergencias": [d.to_dict() for d in self.divergencias],
        }


# ── Helpers de normalização ────────────────────────────────────────────
def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _norm(v: Any) -> Any:
    if isinstance(v, str):
        v = v.strip()
        return v if v != "" else None
    return v


def _is_numeric(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _col_letter(idx0: int) -> str:
    return get_column_letter(idx0 + 1)


def _classificar(v_msc: Any, v_siconfi: Any) -> tuple[str, float | None, float | None]:
    """
    Retorna (classificacao, diferenca_absoluta, diferenca_pct).

    Regra do zero/vazio: célula vazia em um lado e 0 (zero numérico) no
    outro é considerada IGUAL — é só forma distinta de representar "sem
    movimento". Validado nos arquivos reais de Timon/MA 2026.
    """
    n_msc, n_sic = _norm(v_msc), _norm(v_siconfi)

    blank_zero_eq = (
        (n_msc is None and n_sic == 0.0)
        or (n_sic is None and n_msc == 0.0)
    )
    if n_msc == n_sic or blank_zero_eq:
        return "", None, None

    if _is_blank(v_msc) and _is_numeric(n_sic):
        return "AUSENTE_MSC", None, None
    if _is_blank(v_siconfi) and _is_numeric(n_msc):
        return "AUSENTE_SICONFI", None, None

    if _is_numeric(n_msc) and _is_numeric(n_sic):
        diff = round(n_sic - n_msc, 2)
        if abs(diff) < LIMIAR_MINIMA_ABS:
            return "MÍNIMA", diff, None
        base = abs(n_sic) if n_sic != 0 else abs(n_msc)
        pct = abs(diff) / base if base else None
        if pct is None:
            classe = "CRÍTICA"
        elif pct > LIMIAR_CRITICA:
            classe = "CRÍTICA"
        elif pct > LIMIAR_SIGNIFICATIVA:
            classe = "SIGNIFICATIVA"
        elif pct > LIMIAR_MODERADA:
            classe = "MODERADA"
        else:
            classe = "BAIXA"
        return classe, diff, (round(pct * 100, 2) if pct is not None else None)

    return "TEXTO", None, None


def _build_header_map(sheet: xlrd.sheet.Sheet, header_rows: range) -> dict[int, str]:
    headers: dict[int, str] = {}
    for c in range(sheet.ncols):
        parts = []
        for r in header_rows:
            if r < sheet.nrows:
                v = sheet.cell_value(r, c)
                if isinstance(v, str) and v.strip():
                    parts.append(v.strip())
        headers[c] = " / ".join(parts)
    return headers


def _detect_header_rows(sheet: xlrd.sheet.Sheet) -> range:
    rotulo_row = None
    for r in range(sheet.nrows):
        v = sheet.cell_value(r, 0)
        if isinstance(v, str) and v.strip().startswith("Rótulo:"):
            rotulo_row = r
            break
    if rotulo_row is None:
        return range(0, 0)
    return range(rotulo_row + 1, min(rotulo_row + 4, sheet.nrows))


# ── Conversão .xls -> .xlsx preservando layout (LibreOffice headless) ──
def _ensure_xlsx(path: str, workdir: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return path
    if ext != ".xls":
        raise ValueError(f"Formato não suportado: {ext}. Use .xls ou .xlsx.")

    out_dir = os.path.join(workdir, f"conv_{uuid.uuid4().hex[:8]}")
    os.makedirs(out_dir, exist_ok=True)
    cmd = [
        "soffice", "--headless", "--convert-to", "xlsx",
        "--outdir", out_dir, path,
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError(
            f"Falha ao converter {path} para xlsx via LibreOffice: "
            f"{proc.stderr or proc.stdout}"
        )
    base = os.path.splitext(os.path.basename(path))[0]
    converted = os.path.join(out_dir, f"{base}.xlsx")
    if not os.path.exists(converted):
        raise RuntimeError(f"Conversão não gerou o arquivo esperado: {converted}")
    return converted


# ── Núcleo da conciliação ──────────────────────────────────────────────
def conciliar_rreo(
    caminho_msc: str,
    caminho_siconfi: str,
    caminho_saida: str,
    caminho_auditoria: str | None = None,
    workdir: str | None = None,
    ente_nome: str | None = None,
    periodo_nome: str | None = None,
) -> ResultadoConciliacao:
    """
    Compara o rascunho MSC com o arquivo homologado SICONFI, anexo por
    anexo, célula por célula, e gera dois arquivos:

    1. `caminho_saida` — cópia fiel do layout do rascunho com divergências
       destacadas nas células (mesmo padrão do Conciliador RGF).
    2. `caminho_auditoria` — relatório tabular com identidade visual Kora,
       pronto para revisar sem navegar pelos anexos.
    """
    workdir = workdir or tempfile.mkdtemp(prefix="rreo_conc_")
    avisos: list[str] = []

    if caminho_auditoria is None:
        base, _ = os.path.splitext(caminho_saida)
        caminho_auditoria = f"{base}_Auditoria.xlsx"

    dados_msc, _ = _carregar_para_comparacao(caminho_msc)
    dados_siconfi, _ = _carregar_para_comparacao(caminho_siconfi)

    sheets_msc = set(dados_msc.keys())
    sheets_siconfi = set(dados_siconfi.keys())
    sheets_comuns = sorted(sheets_msc & sheets_siconfi, key=lambda s: list(dados_msc.keys()).index(s))

    so_no_msc = sheets_msc - sheets_siconfi
    so_no_siconfi = sheets_siconfi - sheets_msc
    if so_no_msc:
        avisos.append(f"Abas presentes só no rascunho MSC (não comparadas): {', '.join(sorted(so_no_msc))}")
    if so_no_siconfi:
        avisos.append(f"Abas presentes só no SICONFI (não comparadas): {', '.join(sorted(so_no_siconfi))}")

    if ente_nome is None or periodo_nome is None:
        meta_ente, meta_periodo = _extrair_metadados(dados_msc)
        ente_nome = ente_nome or meta_ente
        periodo_nome = periodo_nome or meta_periodo

    xlsx_msc_path = _ensure_xlsx(caminho_msc, workdir)
    shutil.copy(xlsx_msc_path, caminho_saida)
    wb_out = load_workbook(caminho_saida)

    todas_divergencias: list[Divergencia] = []

    for sheet_name in sheets_comuns:
        sheet_msc = dados_msc[sheet_name]
        sheet_siconfi = dados_siconfi[sheet_name]
        headers = sheet_msc["headers"]
        grid_msc = sheet_msc["grid"]
        grid_siconfi = sheet_siconfi["grid"]

        nrows = max(len(grid_msc), len(grid_siconfi))
        ncols = max(sheet_msc["ncols"], sheet_siconfi["ncols"])

        ws_out = wb_out[sheet_name] if sheet_name in wb_out.sheetnames else None
        if ws_out is None:
            avisos.append(f"Aba '{sheet_name}' não encontrada no arquivo de saída convertido — pulada.")
            continue

        for r in range(nrows):
            row_msc = grid_msc[r] if r < len(grid_msc) else []
            row_siconfi = grid_siconfi[r] if r < len(grid_siconfi) else []
            rotulo = None
            if row_msc and isinstance(row_msc[0], str) and row_msc[0].strip():
                rotulo = row_msc[0].strip()
            elif row_siconfi and isinstance(row_siconfi[0], str) and row_siconfi[0].strip():
                rotulo = row_siconfi[0].strip()

            for c in range(ncols):
                v_msc = row_msc[c] if c < len(row_msc) else None
                v_siconfi = row_siconfi[c] if c < len(row_siconfi) else None

                classe, diff, diff_pct = _classificar(v_msc, v_siconfi)
                if not classe:
                    continue

                div = Divergencia(
                    sheet=sheet_name,
                    row=r + 1,
                    col=_col_letter(c),
                    col_idx=c,
                    rotulo=rotulo,
                    cabecalho=headers.get(c) or None,
                    valor_msc=v_msc if not _is_blank(v_msc) else None,
                    valor_siconfi=v_siconfi if not _is_blank(v_siconfi) else None,
                    diferenca=diff,
                    diferenca_pct=diff_pct,
                    classificacao=classe,
                )
                todas_divergencias.append(div)

                cell = ws_out.cell(row=r + 1, column=c + 1)
                cell.fill = DIV_FILLS.get(classe, DIV_FILLS["TEXTO"])
                cell.comment = Comment(_comentario(div), "Kora — Conciliador RREO")

    por_sheet: dict[str, int] = {}
    por_classificacao: dict[str, int] = {}
    for d in todas_divergencias:
        por_sheet[d.sheet] = por_sheet.get(d.sheet, 0) + 1
        por_classificacao[d.classificacao] = por_classificacao.get(d.classificacao, 0) + 1

    wb_out.save(caminho_saida)

    _gerar_relatorio_auditoria(
        caminho_auditoria=caminho_auditoria,
        divergencias=todas_divergencias,
        por_sheet=por_sheet,
        por_classificacao=por_classificacao,
        ente_nome=ente_nome,
        periodo_nome=periodo_nome,
        arquivo_msc_nome=os.path.basename(caminho_msc),
        arquivo_siconfi_nome=os.path.basename(caminho_siconfi),
    )

    return ResultadoConciliacao(
        arquivo_saida=caminho_saida,
        arquivo_auditoria=caminho_auditoria,
        total_divergencias=len(todas_divergencias),
        divergencias=todas_divergencias,
        por_sheet=por_sheet,
        por_classificacao=por_classificacao,
        sheets_comparadas=list(sheets_comuns),
        avisos=avisos,
    )


def _extrair_metadados(dados: dict[str, dict]) -> tuple[str | None, str | None]:
    ente, periodo = None, None
    for sheet in dados.values():
        grid = sheet["grid"]
        for row in grid[:20]:
            if not row:
                continue
            v = row[0]
            if not isinstance(v, str):
                continue
            v = v.strip()
            if v.startswith("Ente:") and ente is None:
                ente = v.replace("Ente:", "").strip()
            elif v.startswith("Período:") and periodo is None:
                periodo = v.replace("Período:", "").strip()
        if ente and periodo:
            break
    return ente, periodo


def _fmt_brl(v: Any) -> str:
    if v is None or v == "":
        return "(vazio)"
    if _is_numeric(v):
        s = f"{v:,.2f}"
        return s.replace(",", "@").replace(".", ",").replace("@", ".")
    return str(v)


def _comentario(d: Divergencia) -> str:
    linhas = [
        f"MSC: {_fmt_brl(d.valor_msc)}",
        f"SICONFI: {_fmt_brl(d.valor_siconfi)}",
    ]
    if d.diferenca is not None:
        linhas.append(f"Dif: {_fmt_brl(d.diferenca)}")
    if d.diferenca_pct is not None:
        linhas.append(f"Dif %: {d.diferenca_pct:.2f}%".replace(".", ","))
    linhas.append(f"Classificação: {d.classificacao}")
    return "\n".join(linhas)


def _carregar_para_comparacao(path: str) -> tuple[dict[str, dict], str]:
    ext = os.path.splitext(path)[1].lower()
    out: dict[str, dict] = {}

    if ext == ".xls":
        book = xlrd.open_workbook(path)
        for sheet_name in book.sheet_names():
            sheet = book.sheet_by_name(sheet_name)
            grid = [
                [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)
            ]
            headers = _build_header_map(sheet, _detect_header_rows(sheet))
            out[sheet_name] = {"grid": grid, "ncols": sheet.ncols, "headers": headers}
        return out, "xls"

    if ext == ".xlsx":
        wb = load_workbook(path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            grid = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
            ncols = ws.max_column or 0
            headers: dict[int, str] = {}
            rotulo_row = None
            for i, row in enumerate(grid[:30]):
                if row and isinstance(row[0], str) and row[0].strip().startswith("Rótulo:"):
                    rotulo_row = i
                    break
            if rotulo_row is not None:
                for c in range(ncols):
                    parts = []
                    for r in range(rotulo_row + 1, min(rotulo_row + 4, len(grid))):
                        v = grid[r][c] if c < len(grid[r]) else None
                        if isinstance(v, str) and v.strip():
                            parts.append(v.strip())
                    headers[c] = " / ".join(parts)
            out[sheet_name] = {"grid": grid, "ncols": ncols, "headers": headers}
        wb.close()
        return out, "xlsx"

    raise ValueError(f"Formato não suportado: {ext}. Use .xls ou .xlsx.")


# ── Relatório de auditoria tabular — identidade visual Kora ────────────
def _gerar_relatorio_auditoria(
    caminho_auditoria: str,
    divergencias: list,
    por_sheet: dict[str, int],
    por_classificacao: dict[str, int],
    ente_nome: str | None,
    periodo_nome: str | None,
    arquivo_msc_nome: str,
    arquivo_siconfi_nome: str,
) -> None:
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_div = wb.create_sheet("Divergências")

    _montar_aba_resumo(ws_resumo, divergencias, por_sheet, por_classificacao,
                       ente_nome, periodo_nome, arquivo_msc_nome, arquivo_siconfi_nome)
    _montar_aba_divergencias(ws_div, divergencias, ente_nome, periodo_nome)

    wb.save(caminho_auditoria)


def _faixa_cabecalho_kora(ws, titulo: str, subtitulo: str | None, largura_colunas: int) -> int:
    last_col_letter = get_column_letter(largura_colunas)

    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = "KORA"
    c.font = Font(name=KORA_FONT_FAMILY, size=18, bold=True, color=KORA_DOURADO)
    c.fill = PatternFill("solid", fgColor=KORA_GRAFITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_col_letter}2")
    c = ws["A2"]
    c.value = titulo
    c.font = Font(name=KORA_FONT_FAMILY, size=12, bold=True, color=KORA_BRANCO)
    c.fill = PatternFill("solid", fgColor=KORA_GRAFITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    next_row = 3
    if subtitulo:
        ws.merge_cells(f"A3:{last_col_letter}3")
        c = ws["A3"]
        c.value = subtitulo
        c.font = Font(name=KORA_FONT_FAMILY, size=9, color=KORA_DOURADO_CLARO)
        c.fill = PatternFill("solid", fgColor=KORA_GRAFITE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[3].height = 18
        next_row = 4

    for col in range(1, largura_colunas + 1):
        ws.cell(row=next_row, column=col).fill = PatternFill("solid", fgColor=KORA_DOURADO)
    ws.row_dimensions[next_row].height = 4
    return next_row + 2


def _montar_aba_resumo(
    ws,
    divergencias: list,
    por_sheet: dict[str, int],
    por_classificacao: dict[str, int],
    ente_nome: str | None,
    periodo_nome: str | None,
    arquivo_msc_nome: str,
    arquivo_siconfi_nome: str,
) -> None:
    ws.sheet_view.showGridLines = False
    larguras = [34, 22, 18]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    subtitulo_partes = [p for p in [ente_nome, periodo_nome] if p]
    subtitulo = " · ".join(subtitulo_partes) if subtitulo_partes else None
    r = _faixa_cabecalho_kora(ws, "Auditoria de Conciliação RREO", subtitulo, largura_colunas=3)

    label_font = Font(name=KORA_FONT_FAMILY, size=9, color=KORA_CINZA_TEXTO)

    def linha_meta(label: str, valor: str, linha: int):
        ws.cell(row=linha, column=1, value=label).font = label_font
        ws.cell(row=linha, column=2, value=valor).font = Font(name=KORA_FONT_FAMILY, size=9, color="FF333333")
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=3)

    linha_meta("Arquivo rascunho (MSC)", arquivo_msc_nome, r)
    linha_meta("Arquivo homologado (SICONFI)", arquivo_siconfi_nome, r + 1)
    r += 3

    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=3)
    cell = ws.cell(row=r, column=1, value=f"{len(divergencias)} divergência(s) identificada(s)")
    cell.font = Font(name=KORA_FONT_FAMILY, size=15, bold=True, color=KORA_GRAFITE)
    cell.fill = PatternFill("solid", fgColor=KORA_MARFIM)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 30
    border_dourado = Border(
        top=Side(style="medium", color=KORA_DOURADO),
        bottom=Side(style="medium", color=KORA_DOURADO),
    )
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = border_dourado
        ws.cell(row=r + 1, column=col).border = Border(bottom=Side(style="medium", color=KORA_DOURADO))
    r += 3

    r = _tabela_secao(ws, r, "Divergências por anexo",
                      sorted(por_sheet.items()), col_label="Anexo")
    r += 1
    ordem_severidade = ["CRÍTICA", "SIGNIFICATIVA", "MODERADA", "BAIXA", "MÍNIMA",
                        "AUSENTE_MSC", "AUSENTE_SICONFI", "TEXTO"]
    itens_classificacao = [(k, por_classificacao[k]) for k in ordem_severidade if k in por_classificacao]
    r = _tabela_secao(ws, r, "Divergências por classificação", itens_classificacao,
                      col_label="Classificação", badge=True)

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    nota = ws.cell(row=r, column=1, value=(
        "Conciliação gerada automaticamente — comparação determinística, célula a célula, "
        "sem intervenção de IA na apuração dos valores. Ver aba \"Divergências\" para o detalhe completo."
    ))
    nota.font = Font(name=KORA_FONT_FAMILY, size=8, italic=True, color=KORA_CINZA_TEXTO)
    nota.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 26

    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _tabela_secao(ws, start_row: int, titulo: str, itens: list,
                  col_label: str, badge: bool = False) -> int:
    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c = ws.cell(row=r, column=1, value=titulo.upper())
    c.font = Font(name=KORA_FONT_FAMILY, size=9, bold=True, color="FFA87A1F")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    thin = Side(style="thin", color=KORA_CINZA_BORDA)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=r, column=1, value=col_label).font = Font(name=KORA_FONT_FAMILY, size=9, bold=True, color=KORA_BRANCO)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=KORA_GRAFITE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=3, value="Qtd.").font = Font(name=KORA_FONT_FAMILY, size=9, bold=True, color=KORA_BRANCO)
    ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=KORA_GRAFITE)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = border
    r += 1

    for i, (nome, qtd) in enumerate(itens):
        fill = KORA_MARFIM if i % 2 == 0 else KORA_BRANCO
        nome_display = nome.replace("RREO-", "") if nome.startswith("RREO-") else nome
        cell_label = ws.cell(row=r, column=1, value=nome_display)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell_qtd = ws.cell(row=r, column=3, value=qtd)
        cell_qtd.alignment = Alignment(horizontal="center")
        cell_qtd.font = Font(name=KORA_FONT_FAMILY, size=9, bold=True)
        if badge and nome in CLASS_BADGE_FILLS:
            cell_label.fill = PatternFill("solid", fgColor=CLASS_BADGE_FILLS[nome])
            cell_label.font = Font(name=KORA_FONT_FAMILY, size=9, bold=True, color=CLASS_BADGE_TEXT[nome])
            cell_qtd.fill = PatternFill("solid", fgColor=CLASS_BADGE_FILLS[nome])
        else:
            cell_label.fill = PatternFill("solid", fgColor=fill)
            cell_label.font = Font(name=KORA_FONT_FAMILY, size=9, color="FF333333")
            cell_qtd.fill = PatternFill("solid", fgColor=fill)
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = border
        r += 1
    return r


def _montar_aba_divergencias(ws, divergencias: list,
                              ente_nome: str | None, periodo_nome: str | None) -> None:
    ws.sheet_view.showGridLines = False
    colunas = ["Anexo", "Célula", "Linha (Rótulo)", "Coluna (Cabeçalho)",
               "Valor MSC", "Valor SICONFI", "Diferença (R$)", "Diferença (%)", "Classificação"]
    larguras = [16, 9, 46, 32, 18, 18, 16, 14, 20]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    subtitulo_partes = [p for p in [ente_nome, periodo_nome] if p]
    subtitulo = " · ".join(subtitulo_partes) if subtitulo_partes else None
    r = _faixa_cabecalho_kora(ws, "Divergências identificadas", subtitulo, largura_colunas=len(colunas))

    header_row = r
    thin = Side(style="thin", color=KORA_CINZA_BORDA)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, titulo in enumerate(colunas, start=1):
        c = ws.cell(row=header_row, column=i, value=titulo)
        c.font = Font(name=KORA_FONT_FAMILY, size=9, bold=True, color=KORA_BRANCO)
        c.fill = PatternFill("solid", fgColor=KORA_GRAFITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[header_row].height = 30

    r = header_row + 1
    for i, d in enumerate(divergencias):
        nome_anexo = d.sheet.replace("RREO-", "") if d.sheet.startswith("RREO-") else d.sheet
        valores = [
            nome_anexo,
            f"{d.col}{d.row}",
            d.rotulo or "",
            d.cabecalho or "",
            d.valor_msc if d.valor_msc is not None else "(vazio)",
            d.valor_siconfi if d.valor_siconfi is not None else "(vazio)",
            d.diferenca,
            (d.diferenca_pct / 100 if d.diferenca_pct is not None else None),
            d.classificacao,
        ]
        base_fill = KORA_MARFIM if i % 2 == 0 else KORA_BRANCO
        for col, valor in enumerate(valores, start=1):
            cell = ws.cell(row=r, column=col, value=valor)
            cell.border = border
            cell.font = Font(name=KORA_FONT_FAMILY, size=9, color="FF2A2A2A")
            if col == 9:
                cell.fill = PatternFill("solid", fgColor=CLASS_BADGE_FILLS.get(d.classificacao, "FFEAEAEA"))
                cell.font = Font(name=KORA_FONT_FAMILY, size=9, bold=True,
                                 color=CLASS_BADGE_TEXT.get(d.classificacao, "FF555555"))
                cell.alignment = Alignment(horizontal="center")
                continue
            cell.fill = PatternFill("solid", fgColor=base_fill)
            if col in (5, 6, 7) and isinstance(valor, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col == 8 and isinstance(valor, (int, float)):
                cell.number_format = "0.00%"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=(col in (3, 4)))
        r += 1

    ws.freeze_panes = f"A{header_row + 1}"
    last_col = get_column_letter(len(colunas))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{r - 1}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = False


RESULTADO_SCHEMA = {
    "arquivo_saida": "str — caminho do .xlsx gerado (planilha com highlights)",
    "arquivo_auditoria": "str — caminho do .xlsx de auditoria Kora",
    "total_divergencias": "int",
    "por_sheet": "dict[str, int] — divergências por anexo",
    "por_classificacao": "dict[str, int] — divergências por severidade",
    "sheets_comparadas": "list[str] — anexos presentes em ambos os arquivos",
    "avisos": "list[str] — abas ausentes em um dos lados, etc.",
    "divergencias": "list[dict] — cada item com sheet/celula/rotulo/cabecalho/"
                    "valor_msc/valor_siconfi/diferenca/diferenca_pct/classificacao",
}


if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 4:
        print("Uso: python conciliador_rreo.py <rascunho_msc> <siconfi> <saida.xlsx>")
        sys.exit(1)

    resultado = conciliar_rreo(sys.argv[1], sys.argv[2], sys.argv[3])
    print(json.dumps(resultado.to_dict(), ensure_ascii=False, indent=2, default=str)[:3000])
    print(f"\n>>> {resultado.total_divergencias} divergências.")
    print(f">>> Planilha destacada: {resultado.arquivo_saida}")
    print(f">>> Relatório de auditoria: {resultado.arquivo_auditoria}")
