"""
conciliador.py
--------------
Algoritmo de conciliação RGF Simplificado.

Recebe dois arquivos (rascunho MSC e homologado SICONFI), ambos .xls ou .xlsx,
e gera um .xlsx que é cópia fiel do rascunho com as divergências marcadas
nas próprias células — mesmas cores, merges, larguras, alturas — mais popup
de comentário com valores detalhados.

Classificação de divergências:
  CRÍTICA        diff > 20 %
  SIGNIFICATIVA  5 % – 20 %
  MODERADA       1 % – 5 %
  BAIXA          < 1 %
  MÍNIMA         diff abs < R$ 0,01
  AUSENTE        valor presente em um arquivo, ausente no outro
"""

import os
import numpy as np
import pandas as pd
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment


# ── Palette ──────────────────────────────────────────────────────────
DIV_FILLS = {
    "CRÍTICA":       PatternFill("solid", fgColor="FFFF3333"),
    "SIGNIFICATIVA": PatternFill("solid", fgColor="FFFFAA00"),
    "MODERADA":      PatternFill("solid", fgColor="FFFFFF44"),
    "BAIXA":         PatternFill("solid", fgColor="FF90EE90"),
    "MÍNIMA":        PatternFill("solid", fgColor="FFD4F1D4"),
    "AUSENTE":       PatternFill("solid", fgColor="FFADD8E6"),
}
DIV_FONT_COLOR = {
    "CRÍTICA":       "FFFFFFFF",
    "SIGNIFICATIVA": "FF000000",
    "MODERADA":      "FF000000",
    "BAIXA":         "FF000000",
    "MÍNIMA":        "FF000000",
    "AUSENTE":       "FF003366",
}


# ── Helpers ──────────────────────────────────────────────────────────
def _rgb_hex(r, g, b) -> str:
    return f"FF{r:02X}{g:02X}{b:02X}"


def _to_num(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    try:
        return float(str(val).replace(",", ".").strip())
    except Exception:
        return None


def _classify(diff: float, msc: float | None, sic: float | None) -> str | None:
    if msc is None and sic is not None:
        return "AUSENTE"
    if sic is None and msc is not None:
        return "AUSENTE"
    if abs(diff) < 0.005:
        return None
    if abs(diff) < 0.01:
        return "MÍNIMA"
    if sic == 0:
        return "CRÍTICA"
    pct = abs(diff) / abs(sic) * 100
    if pct < 1:
        return "BAIXA"
    if pct < 5:
        return "MODERADA"
    if pct < 20:
        return "SIGNIFICATIVA"
    return "CRÍTICA"


def _xlrd_bg(wb_x, xf_idx) -> str | None:
    xf = wb_x.xf_list[xf_idx]
    rgb = wb_x.colour_map.get(xf.background.pattern_colour_index)
    return _rgb_hex(*rgb) if rgb else None


def _xlrd_font(wb_x, xf_idx) -> dict:
    xf = wb_x.xf_list[xf_idx]
    fnt = wb_x.font_list[xf.font_index]
    rgb = wb_x.colour_map.get(fnt.colour_index)
    return dict(
        name=fnt.name or "Arial",
        size=fnt.height / 20,
        bold=bool(fnt.bold),
        italic=bool(fnt.italic),
        color=_rgb_hex(*rgb) if rgb else "FF000000",
    )


def _xlrd_align(wb_x, xf_idx) -> dict:
    a = wb_x.xf_list[xf_idx].alignment
    H = {0: "general", 1: "left", 2: "center", 3: "right",
         4: "fill", 5: "justify", 6: "centerContinuous"}
    V = {0: "top", 1: "center", 2: "bottom", 3: "justify"}
    return dict(
        horizontal=H.get(a.hor_align, "general"),
        vertical=V.get(a.vert_align, "bottom"),
        wrap=bool(a.text_wrapped),
    )


def _xlrd_numfmt(wb_x, xf_idx) -> str:
    xf = wb_x.xf_list[xf_idx]
    fmt = wb_x.format_map.get(xf.format_key)
    return fmt.format_str if fmt else "General"


def _xlrd_border(wb_x, xf_idx) -> Border:
    b = wb_x.xf_list[xf_idx].border
    STYLES = {
        0: None, 1: "thin", 2: "medium", 3: "dashed", 4: "dotted",
        5: "thick", 6: "double", 7: "hair", 8: "mediumDashed",
        9: "dashDot", 10: "mediumDashDot", 11: "dashDotDot",
        12: "mediumDashDotDot", 13: "slantDashDot",
    }

    def side(ls, ci):
        st = STYLES.get(ls)
        if not st:
            return Side(style=None)
        rgb = wb_x.colour_map.get(ci)
        return Side(style=st, color=_rgb_hex(*rgb) if rgb else "FF000000")

    return Border(
        left=side(b.left_line_style, b.left_colour_index),
        right=side(b.right_line_style, b.right_colour_index),
        top=side(b.top_line_style, b.top_colour_index),
        bottom=side(b.bottom_line_style, b.bottom_colour_index),
    )


# ── openpyxl formatting helpers (fallback para .xls BIFF8-obfuscated) ─
def _opxl_bg(cell) -> str | None:
    fill = cell.fill
    if fill and fill.fill_type == "solid" and fill.fgColor:
        rgb = fill.fgColor.rgb  # "FFRRGGBB"
        if rgb and rgb not in ("00000000", "FF000000"):
            return rgb
    return None


def _opxl_font(cell) -> dict:
    f = cell.font
    color = "FF000000"
    if f and f.color and f.color.type == "rgb":
        color = f.color.rgb
    return dict(
        name=(f.name if f and f.name else "Arial"),
        size=(f.size if f and f.size else 10),
        bold=bool(f.bold if f else False),
        italic=bool(f.italic if f else False),
        color=color,
    )


def _opxl_align(cell) -> dict:
    a = cell.alignment
    return dict(
        horizontal=(a.horizontal or "general") if a else "general",
        vertical=(a.vertical or "bottom") if a else "bottom",
        wrap=bool(a.wrap_text if a else False),
    )


def _opxl_numfmt(cell) -> str:
    return cell.number_format or "General"


def _opxl_border(cell) -> Border:
    b = cell.border
    if not b:
        return Border()
    STYLES = {
        None: None, "thin": "thin", "medium": "medium", "thick": "thick",
        "dashed": "dashed", "dotted": "dotted", "double": "double",
        "hair": "hair", "mediumDashed": "mediumDashed",
        "dashDot": "dashDot", "mediumDashDot": "mediumDashDot",
        "dashDotDot": "dashDotDot", "mediumDashDotDot": "mediumDashDotDot",
        "slantDashDot": "slantDashDot",
    }

    def _side(s):
        if not s or not s.border_style:
            return Side(style=None)
        st = STYLES.get(s.border_style, "thin")
        color = "FF000000"
        if s.color and s.color.type == "rgb":
            color = s.color.rgb
        return Side(style=st, color=color)

    return Border(
        left=_side(b.left), right=_side(b.right),
        top=_side(b.top), bottom=_side(b.bottom),
    )


def _open_xls(path: str):
    """Open .xls with xlrd (formatting_info=True).
    Falls back to LibreOffice→xlsx→openpyxl when BIFF8 XOR obfuscation is detected."""
    with open(path, "rb") as f:
        raw = f.read()
    try:
        return xlrd.open_workbook(file_contents=raw, formatting_info=True)
    except Exception as e:
        if "encrypted" not in str(e).lower():
            raise
    # BIFF8 XOR obfuscation — convert to xlsx (OOXML strips the obfuscation)
    import tempfile, subprocess
    with tempfile.TemporaryDirectory() as tmp:
        src = os.path.join(tmp, os.path.basename(path))
        with open(src, "wb") as f:
            f.write(raw)
        proc = subprocess.run(
            ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", tmp, src],
            capture_output=True, text=True, timeout=120,
        )
        if proc.returncode != 0:
            raise RuntimeError(f"LibreOffice falhou ao converter arquivo: {proc.stderr or proc.stdout}")
        base = os.path.splitext(os.path.basename(path))[0]
        converted = os.path.join(tmp, f"{base}.xlsx")
        if not os.path.exists(converted):
            raise RuntimeError(f"LibreOffice não gerou o arquivo esperado: {converted}")
        return load_workbook(converted)  # openpyxl.Workbook


# ── Legend sheet ─────────────────────────────────────────────────────
def _add_legend(wb_out: Workbook):
    ws = wb_out.create_sheet("🔑 Legenda")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 50

    th = Side(style="thin", color="FF888888")
    bdr = Border(left=th, right=th, top=th, bottom=th)

    ws.merge_cells("A1:D1")
    ws["A1"] = "LEGENDA – MARCAÇÕES DE DIVERGÊNCIA  |  RGF Simplificado"
    ws["A1"].font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="FF1B2A4A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:D2")
    ws["A2"] = (
        "Células coloridas = divergência entre o Rascunho MSC e o SICONFI homologado. "
        "Passe o cursor sobre a célula para ver o popup: valor MSC, valor SICONFI, diferença e percentual."
    )
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="FF333333")
    ws["A2"].fill = PatternFill("solid", fgColor="FFF0F4FA")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32

    for c, h in enumerate(["Cor", "Classificação", "Critério", "Ação Recomendada"], 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF2D4A8A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
    ws.row_dimensions[3].height = 20

    rows = [
        ("FFFF3333", "FFFFFFFF", "CRÍTICA (> 20%)",        "Diferença > 20% do valor SICONFI",     "Correção imediata obrigatória antes de nova transmissão"),
        ("FFFFAA00", "FF000000", "SIGNIFICATIVA (5–20%)",  "Diferença entre 5% e 20%",             "Verificar lançamentos e validar com o contador responsável"),
        ("FFFFFF44", "FF000000", "MODERADA (1–5%)",        "Diferença entre 1% e 5%",              "Conferir memória de cálculo, competências e deduções"),
        ("FF90EE90", "FF000000", "BAIXA (< 1%)",           "Diferença inferior a 1%",              "Revisar arredondamentos e regras de conversão"),
        ("FFD4F1D4", "FF000000", "MÍNIMA (centavos)",      "Diferença < R$ 0,01",                  "Verificar precisão decimal na exportação"),
        ("FFADD8E6", "FF003366", "AUSENTE",                "Presente em um arquivo, ausente no outro", "Verificar preenchimento e completude do envio"),
    ]
    for i, (bg, fc, cls, crit, acao) in enumerate(rows):
        r = 4 + i
        for c, (val, align) in enumerate([
            ("■", "center"), (cls, "left"), (crit, "center"), (acao, "left")
        ], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = PatternFill("solid", fgColor=(bg if c < 4 else "FFFAFAFA"))
            cell.font = Font(
                name="Arial", size=(14 if c == 1 else 9),
                bold=(c <= 2), color=(fc if c <= 3 else "FF222222"),
            )
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = bdr
        ws.row_dimensions[r].height = 22


# ── Main entry point ─────────────────────────────────────────────────
def conciliar(rascunho_path: str, homologado_path: str, output_path: str) -> dict:
    """
    Gera um XLSX que é cópia fiel do rascunho com células divergentes marcadas.

    Returns a stats dict:
        {
            "total_divergencias": int,
            "por_severidade": {"CRÍTICA": n, ...},
            "por_anexo": {"RGF-Anexo 01": n, ...},
        }
    """
    r_ext = os.path.splitext(rascunho_path)[1].lower()
    h_ext = os.path.splitext(homologado_path)[1].lower()

    # Load rascunho — xlrd normal, openpyxl como fallback para BIFF8 obfuscated
    r_wb_x = _open_xls(rascunho_path)
    _using_openpyxl = not isinstance(r_wb_x, xlrd.Book)

    # Load homologado with pandas (only values needed)
    engine_h = "xlrd" if h_ext == ".xls" else "openpyxl"
    try:
        s_df_all = pd.read_excel(homologado_path, engine=engine_h,
                                  sheet_name=None, header=None)
    except Exception as e:
        if "encrypted" not in str(e).lower() or h_ext != ".xls":
            raise
        import io, tempfile, subprocess
        with open(homologado_path, "rb") as f:
            raw_h = f.read()
        with tempfile.TemporaryDirectory() as tmp:
            src = os.path.join(tmp, os.path.basename(homologado_path))
            with open(src, "wb") as f:
                f.write(raw_h)
            proc = subprocess.run(
                ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", tmp, src],
                capture_output=True, text=True, timeout=120,
            )
            if proc.returncode != 0:
                raise RuntimeError(f"LibreOffice falhou: {proc.stderr or proc.stdout}")
            base = os.path.splitext(os.path.basename(homologado_path))[0]
            converted = os.path.join(tmp, f"{base}.xlsx")
            with open(converted, "rb") as f:
                conv_raw_h = f.read()
        s_df_all = pd.read_excel(io.BytesIO(conv_raw_h), engine="openpyxl",
                                  sheet_name=None, header=None)

    sheet_names = r_wb_x.sheetnames if _using_openpyxl else r_wb_x.sheet_names()

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    stats: dict = {
        "total_divergencias": 0,
        "por_severidade": {k: 0 for k in DIV_FILLS},
        "por_anexo": {},
    }

    for sheet_name in sheet_names:
        s_df = s_df_all.get(sheet_name)
        ws = wb_out.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        # Build SICONFI numeric lookup (same for both paths)
        sic_lookup: dict[tuple[int, int], float] = {}
        if s_df is not None:
            for ri in range(len(s_df)):
                for ci in range(len(s_df.columns)):
                    v = _to_num(s_df.iloc[ri, ci])
                    if v is not None:
                        sic_lookup[(ri, ci)] = v

        sheet_divs = 0

        if _using_openpyxl:
            # ── openpyxl path (rascunho BIFF8-obfuscated, converted via LibreOffice) ──
            r_ws = r_wb_x[sheet_name]
            ncols = r_ws.max_column or 0
            nrows = r_ws.max_row or 0

            for c_idx in range(ncols):
                col_letter = get_column_letter(c_idx + 1)
                dim = r_ws.column_dimensions.get(col_letter)
                if dim and dim.width:
                    ws.column_dimensions[col_letter].width = dim.width

            for r_idx in range(nrows):
                dim = r_ws.row_dimensions.get(r_idx + 1)
                if dim and dim.height:
                    ws.row_dimensions[r_idx + 1].height = dim.height

            slave_cells: set[tuple[int, int]] = set()
            for mc in r_ws.merged_cells.ranges:
                r1 = mc.min_row - 1
                r2 = mc.max_row
                c1 = mc.min_col - 1
                c2 = mc.max_col
                for rr in range(r1, r2):
                    for cc in range(c1, c2):
                        if rr != r1 or cc != c1:
                            slave_cells.add((rr, cc))
            for mc in r_ws.merged_cells.ranges:
                ws.merge_cells(str(mc))

            for r_idx in range(nrows):
                for c_idx in range(ncols):
                    if (r_idx, c_idx) in slave_cells:
                        continue

                    cell = r_ws.cell(row=r_idx + 1, column=c_idx + 1)
                    ox = ws.cell(row=r_idx + 1, column=c_idx + 1)

                    val = cell.value if cell.value is not None else ""
                    ox.value = val

                    bg = _opxl_bg(cell)
                    if bg:
                        ox.fill = PatternFill("solid", fgColor=bg)

                    fi = _opxl_font(cell)
                    ox.font = Font(name=fi["name"], size=fi["size"],
                                   bold=fi["bold"], italic=fi["italic"], color=fi["color"])

                    ai = _opxl_align(cell)
                    ox.alignment = Alignment(horizontal=ai["horizontal"],
                                             vertical=ai["vertical"], wrap_text=ai["wrap"])

                    fmt = _opxl_numfmt(cell)
                    if fmt and fmt != "General":
                        ox.number_format = fmt

                    ox.border = _opxl_border(cell)

                    msc_num = _to_num(val)
                    sic_num = sic_lookup.get((r_idx, c_idx))

                    if msc_num is not None or sic_num is not None:
                        diff = (msc_num or 0) - (sic_num or 0)
                        key = _classify(diff, msc_num, sic_num)
                        if key:
                            sheet_divs += 1
                            stats["total_divergencias"] += 1
                            stats["por_severidade"][key] += 1

                            ox.fill = DIV_FILLS[key]
                            ox.font = Font(name=fi["name"], size=fi["size"],
                                           bold=True, italic=fi["italic"],
                                           color=DIV_FONT_COLOR[key])

                            msc_s = f"{msc_num:,.2f}" if msc_num is not None else "ausente"
                            sic_s = f"{sic_num:,.2f}" if sic_num is not None else "ausente"
                            diff_s = f"{diff:+,.2f}"
                            pct_s = (
                                f"{diff / sic_num * 100:+.2f}%"
                                if sic_num and sic_num != 0 else "N/A"
                            )
                            txt = (
                                f"⚠ DIVERGÊNCIA [{key}]\n"
                                f"MSC Rascunho:  {msc_s}\n"
                                f"SICONFI:       {sic_s}\n"
                                f"Diferença:     {diff_s} ({pct_s})"
                            )
                            cmt = Comment(txt, "Kora Audit")
                            cmt.width = 290
                            cmt.height = 88
                            ox.comment = cmt

        else:
            # ── xlrd path (rascunho normal, sem obfuscação) ──────────────
            r_ws_x = r_wb_x.sheet_by_name(sheet_name)

            for c_idx in range(r_ws_x.ncols):
                ci = r_ws_x.colinfo_map.get(c_idx)
                if ci:
                    ws.column_dimensions[get_column_letter(c_idx + 1)].width = max(ci.width / 256.0, 2)

            for r_idx in range(r_ws_x.nrows):
                ri = r_ws_x.rowinfo_map.get(r_idx)
                if ri:
                    ws.row_dimensions[r_idx + 1].height = ri.height / 20.0

            slave_cells: set[tuple[int, int]] = set()
            for mc in r_ws_x.merged_cells:
                r1, r2, c1, c2 = mc
                for rr in range(r1, r2):
                    for cc in range(c1, c2):
                        if rr != r1 or cc != c1:
                            slave_cells.add((rr, cc))

            for mc in r_ws_x.merged_cells:
                r1, r2, c1, c2 = mc
                if r2 - r1 > 0 or c2 - c1 > 0:
                    ws.merge_cells(
                        start_row=r1 + 1, start_column=c1 + 1,
                        end_row=r2, end_column=c2,
                    )

            for r_idx in range(r_ws_x.nrows):
                for c_idx in range(r_ws_x.ncols):
                    if (r_idx, c_idx) in slave_cells:
                        continue

                    cell_x = r_ws_x.cell(r_idx, c_idx)
                    xf_idx = r_ws_x.cell_xf_index(r_idx, c_idx)
                    ox = ws.cell(row=r_idx + 1, column=c_idx + 1)

                    val = cell_x.value if cell_x.ctype not in (0, 6) else None
                    ox.value = val

                    bg = _xlrd_bg(r_wb_x, xf_idx)
                    if bg:
                        ox.fill = PatternFill("solid", fgColor=bg)

                    fi = _xlrd_font(r_wb_x, xf_idx)
                    ox.font = Font(name=fi["name"], size=fi["size"],
                                   bold=fi["bold"], italic=fi["italic"], color=fi["color"])

                    ai = _xlrd_align(r_wb_x, xf_idx)
                    ox.alignment = Alignment(horizontal=ai["horizontal"],
                                             vertical=ai["vertical"], wrap_text=ai["wrap"])

                    fmt = _xlrd_numfmt(r_wb_x, xf_idx)
                    if fmt and fmt != "General":
                        ox.number_format = fmt

                    ox.border = _xlrd_border(r_wb_x, xf_idx)

                    msc_num = _to_num(val)
                    sic_num = sic_lookup.get((r_idx, c_idx))

                    if msc_num is not None or sic_num is not None:
                        diff = (msc_num or 0) - (sic_num or 0)
                        key = _classify(diff, msc_num, sic_num)
                        if key:
                            sheet_divs += 1
                            stats["total_divergencias"] += 1
                            stats["por_severidade"][key] += 1

                            ox.fill = DIV_FILLS[key]
                            ox.font = Font(name=fi["name"], size=fi["size"],
                                           bold=True, italic=fi["italic"],
                                           color=DIV_FONT_COLOR[key])

                            msc_s = f"{msc_num:,.2f}" if msc_num is not None else "ausente"
                            sic_s = f"{sic_num:,.2f}" if sic_num is not None else "ausente"
                            diff_s = f"{diff:+,.2f}"
                            pct_s = (
                                f"{diff / sic_num * 100:+.2f}%"
                                if sic_num and sic_num != 0 else "N/A"
                            )
                            txt = (
                                f"⚠ DIVERGÊNCIA [{key}]\n"
                                f"MSC Rascunho:  {msc_s}\n"
                                f"SICONFI:       {sic_s}\n"
                                f"Diferença:     {diff_s} ({pct_s})"
                            )
                            cmt = Comment(txt, "Kora Audit")
                            cmt.width = 290
                            cmt.height = 88
                            ox.comment = cmt

        stats["por_anexo"][sheet_name] = sheet_divs

    # Legend first
    _add_legend(wb_out)
    wb_out.move_sheet("🔑 Legenda", offset=-(len(wb_out.sheetnames) - 1))

    wb_out.save(output_path)
    return stats
