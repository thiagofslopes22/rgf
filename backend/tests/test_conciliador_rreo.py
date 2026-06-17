"""
Testes do conciliador_rreo.py — rodar com: pytest tests/test_conciliador_rreo.py -v

Usa os arquivos reais de Timon/MA 2026 (2º bimestre) como fixture de
regressão. O resultado esperado (258 divergências) foi validado
manualmente célula a célula antes de existir este módulo — qualquer
mudança futura no algoritmo que quebre esse número precisa ser
investigada antes de subir para produção.

Para rodar os testes de regressão (test_total_divergencias_regressao etc.)
coloque os arquivos reais em:
  backend/tests/fixtures/rascunho_RREO_2026.xls
  backend/tests/fixtures/SICONFI_RREO_1517_BIMESTRAL_2_2026.xls
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from conciliador_rreo import conciliar_rreo, _classificar  # noqa: E402

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")
MSC = os.path.join(FIXTURES, "rascunho_RREO_2026.xls")
SICONFI = os.path.join(FIXTURES, "SICONFI_RREO_1517_BIMESTRAL_2_2026.xls")

FIXTURES_EXIST = os.path.exists(MSC) and os.path.exists(SICONFI)
needs_fixtures = pytest.mark.skipif(not FIXTURES_EXIST, reason="Fixtures de regressão não encontradas em tests/fixtures/")


@pytest.fixture(scope="module")
def resultado(tmp_path_factory):
    out = tmp_path_factory.mktemp("out") / "saida.xlsx"
    return conciliar_rreo(MSC, SICONFI, str(out))


@needs_fixtures
def test_total_divergencias_regressao(resultado):
    """Número de divergências validado manualmente nos arquivos reais."""
    assert resultado.total_divergencias == 258


@needs_fixtures
def test_todos_anexos_comparados(resultado):
    esperado = {
        "RREO-Anexo 01", "RREO-Anexo 02", "RREO-Anexo 03", "RREO-Anexo 04",
        "RREO-Anexo 06", "RREO-Anexo 07", "RREO-Anexo 13", "RREO-Anexo 14",
    }
    assert set(resultado.sheets_comparadas) == esperado


@needs_fixtures
def test_arquivo_saida_existe_e_abre(resultado):
    from openpyxl import load_workbook
    assert os.path.exists(resultado.arquivo_saida)
    wb = load_workbook(resultado.arquivo_saida)
    assert "RREO-Anexo 01" in wb.sheetnames


@needs_fixtures
def test_arquivo_auditoria_existe_e_tem_design_kora(resultado):
    from openpyxl import load_workbook
    assert resultado.arquivo_auditoria is not None
    assert os.path.exists(resultado.arquivo_auditoria)
    wb = load_workbook(resultado.arquivo_auditoria)
    assert "Resumo" in wb.sheetnames
    assert "Divergências" in wb.sheetnames
    ws = wb["Resumo"]
    assert ws["A1"].value == "KORA"
    assert ws["A1"].fill.fgColor.rgb == "FF0E1117"
    assert ws["A1"].font.color.rgb == "FFC9962F"


@needs_fixtures
def test_aba_divergencias_tem_todas_as_linhas(resultado):
    from openpyxl import load_workbook
    wb = load_workbook(resultado.arquivo_auditoria)
    ws = wb["Divergências"]
    assert ws.max_row >= resultado.total_divergencias


@needs_fixtures
def test_celula_conhecida_b21_anexo01(resultado):
    """RECEITAS (EXCETO INTRA-ORÇAMENTÁRIAS) (I) — PREVISÃO INICIAL."""
    achou = [
        d for d in resultado.divergencias
        if d.sheet == "RREO-Anexo 01" and d.col == "B" and d.row == 21
    ]
    assert len(achou) == 1
    d = achou[0]
    assert d.valor_msc == 1197657394.0
    assert d.valor_siconfi == 1211147394.0
    assert d.diferenca == 13490000.0


@needs_fixtures
def test_avisos_quando_falta_anexo(tmp_path):
    out = tmp_path / "saida2.xlsx"
    r = conciliar_rreo(MSC, SICONFI, str(out))
    assert isinstance(r.avisos, list)


# ── Testes unitários (não precisam de fixtures) ────────────────────────

def test_sem_falso_positivo_zero_vs_vazio():
    """Vazio no MSC + 0,00 no SICONFI não deve gerar divergência."""
    classe, diff, pct = _classificar("", 0.0)
    assert classe == ""
    classe, diff, pct = _classificar(0.0, "")
    assert classe == ""


def test_valor_identico_nao_diverge():
    classe, _, _ = _classificar(100.0, 100.0)
    assert classe == ""


def test_ausente_no_msc_detectado():
    classe, diff, pct = _classificar("", 13490000.0)
    assert classe == "AUSENTE_MSC"


def test_ausente_no_siconfi_detectado():
    classe, diff, pct = _classificar(13490000.0, "")
    assert classe == "AUSENTE_SICONFI"


def test_severidade_critica():
    classe, diff, pct = _classificar(100.0, 200.0)  # 50% de diferença
    assert classe == "CRÍTICA"
    assert pct == 50.0


def test_severidade_significativa():
    classe, diff, pct = _classificar(100.0, 110.0)  # ~9% sobre o homologado
    assert classe == "SIGNIFICATIVA"


def test_severidade_moderada():
    classe, diff, pct = _classificar(100.0, 102.0)  # ~1.96%
    assert classe == "MODERADA"


def test_severidade_baixa():
    classe, diff, pct = _classificar(1000.0, 1005.0)  # 0.5%
    assert classe == "BAIXA"


def test_minima_centavos():
    classe, diff, pct = _classificar(100.00, 100.005)
    assert classe == "MÍNIMA"


def test_diferenca_negativa_calculada_corretamente():
    classe, diff, pct = _classificar(200.0, 100.0)
    assert diff == -100.0
    assert classe == "CRÍTICA"


def test_none_vs_none_nao_diverge():
    classe, _, _ = _classificar(None, None)
    assert classe == ""


def test_texto_divergente():
    classe, _, _ = _classificar("Nota publicada A", "Nota publicada B")
    assert classe == "TEXTO"


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
